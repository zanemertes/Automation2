
os.chdir('C:\\Users\\zane.mertes\\PycharmProjects\\IRB') - change working directory
openpyxl.load_workbook('BS.xlsx') - open an excelfile

#Excel functions
#im kurzen

import openpyxl
import os

class IRBmethods(object):
    os.chdir('C:\\Users\\zane.mertes\\PycharmProjects\\IRB')

    def __init__(self):
        self.BS = openpyxl.load_workbook('BS.xlsx')
        self.sheet = self.BS['Sheet']

    def line_find(self, compare_with: str, column_no: int, start_at: int = 1, end_at: int = 1000) -> int:
          for i in range(start_at, end_at):
            v = self.sheet.cell(i, column_no).value
            if compare_with in v:
                print("Column number {0}: {1}".format(i, v))

    def column_find(self, compare_with: str, row_no: int=1, start_at: int = 1, end_at: int = 1000) -> int:
        for i in range(start_at, end_at):
            v = self.sheet.cell(row_no, i).value
            if compare_with in v:
                print("Column number {0}: {1}".format(i, v))


    def corr_number (self, ref_no: str, row_name: str):
        row = self.line_find(ref_no, 1)
        col = self.column_find(row_name)
        my_number = self.sheet.cell(row, col).value
        print(my_number)

#IRB code until line 250
import openpyxl
import os
from itertools import chain


def line_find(compare_with: str, column_no: int, start_at: int = 1, end_at: int = 1000) -> int:
    BS = openpyxl.load_workbook('BS.xlsx')
    sheet = BS['Sheet']

    for i in range(start_at, end_at):
        v = sheet.cell(i, column_no).value
        if compare_with in v:
            print("Row number {0}: {1}".format(i, v))
            return i


def column_find(compare_with: str, row_no: int=1, start_at: int = 1, end_at: int = 1000) -> int:
    BS = openpyxl.load_workbook('BS.xlsx')
    sheet = BS['Sheet']

    for i in range(start_at, end_at):
        v = sheet.cell(row_no, i).value
        if compare_with in v:
            print("Column number {0}: {1}".format(i, v))
            return i


def corr_number (ref_no: str, row_name: str):
    BS = openpyxl.load_workbook('BS.xlsx')
    sheet = BS['Sheet']

    row = line_find(ref_no, 1)
    col = column_find(row_name)
    my_fantastic_number = sheet.cell(row, col).value
    return my_fantastic_number


def MB_line_find(MB_compare_with: str, MB_column_no: int, MB_start_at: int = 1, MB_end_at: int = 1000) -> int:
    MBcp = openpyxl.load_workbook('MBcp.xlsx')
    SCR = MBcp['SCR']

    for i in range(MB_start_at, MB_end_at):
        v = SCR.cell(i, MB_column_no).value
        if v is None:
            print("Whoops. Value is none!")
        else:
            if MB_compare_with in str(v):
                print("Row number {0}: {1}".format(i, v))
                return i


def MB_column_find(MB_compare_with: str, MB_row_no: int=1, MB_start_at: int = 1, MB_end_at: int = 1000) -> int:
    MBcp = openpyxl.load_workbook('MBcp.xlsx')
    SCR = MBcp['SCR']

    for i in range(MB_start_at, MB_end_at):
        v = SCR.cell(MB_row_no, i).value
        if MB_compare_with in v:
            print("Column number {0}: {1}".format(i, v))
            return i


def MB_corr_number (MB_ref_no: str, MB_row_name: str):
    MBcp = openpyxl.load_workbook('MBcp.xlsx')
    SCR = MBcp['SCR']

    MB_row = MB_line_find(MB_ref_no, 1)
    MB_col = MB_column_find(MB_row_name)
    my_fantastic_number = SCR.cell(MB_row, MB_col).value
    return my_fantastic_number

os.chdir('C:\\Users\\zane.mertes\\PycharmProjects\\IRB')


print("Name balance sheet as BS, Marktwertbilanz from previous period as MBpp. "
      "Marktwertbilanz from current period as MBcp, Ausfallrisiko as CP! The report as IRB"
      "Delete empty lines and rows in MBcp. Add necessary information in MBcp SCR. Save numbers as numbers "
      "Saldierte stille Reserven + Lasten I has to be filled manually. Zinsänderungsrisiko mistake in the code. ")

cwd = os.getcwd()
os.chdir('C:\\Users\\zane.mertes\\PycharmProjects\\IRB')

_default_file = 'IRB.xlsx'
filename = input("Which excel file do you want to work on?\nPer default = {0}\n".format(_default_file))
if filename == '':
    filename = _default_file
IRB = openpyxl.load_workbook(filename)
# IRB = openpyxl.load_workbook('IRB_2018_Q1_final.xlsx')

Q = input("Which quarter is it? (Just the number)")

Y = input("Which year is it? ")

D = input("Stichtag? (DD.MM.YYYY)")

first = IRB['1) Inhalt']
first['B4'].value = "Q{0}/{1}".format(Q, Y)
first['B6'].value = D

# Commentare und Graphiken fehlen.

# G
_range = range(6, 16+1)
second = IRB['2) RTF HGB']

for i in _range:
    second['G'+str(i)].value = second['I'+str(i)].value

# F from BS
Kapitalanlagen = corr_number('III. Sonstige Kapitalanlagen', 'Saldo')
Aktien = corr_number('Summe Aktien', 'Saldo')

ZinsTR = Kapitalanlagen - Aktien

second['F6'].value = ZinsTR
second['F11'].value = ZinsTR
second['F12'].value = ZinsTR

second['F9'].value = Aktien

# F from CP
CP = openpyxl.load_workbook('CP.xlsx')

CPT1 = CP['Ausfallrisiko Typ 1']


BAbbb = 0
BAa = 0

for i in range(19, 100):
    if "" == CPT1['C' + str(i)].value or CPT1['C' + str(i)].value is None:
        break
    if "BBB" in CPT1['C' + str(i)].value:
        BAbbb += CPT1['D' + str(i)].value
    print(BAbbb)  # when it is a Verknüpfung in CP, it takes it as a string not as an integer. Change???

second['F7'].value = BAbbb

for i in range(19, 100):
    if "" == CPT1['C' + str(i)].value or CPT1['C' + str(i)].value is None:
        break
    if "A" in CPT1['C' + str(i)].value:
        BAa += CPT1['D' + str(i)].value
    print(BAa)

second['F8'].value = BAa

# Deckungskapital HGB

Eigenkapital = corr_number('A. Eigenkapital', 'Saldo')
second['G33'].value = Eigenkapital

#Solvency II

_third = IRB['3) RTF S II']

_range_chain = range(5, 28 + 1)
_range_chain = chain(_range_chain, range(30, 31+1))
_range_chain = chain(_range_chain, range(36, 38+1))
_range_chain = chain(_range_chain, range(40, 47+1))
_range_chain = chain(_range_chain, range(51, 53+1))

for _i in _range_chain:
    _third['E'+str(_i)].value = _third['G'+str(_i)].value

    zins_r = MB_corr_number('Zinsrisiko', 'SCR netto')
    _third['G5'].value = zins_r



_third = IRB['3) RTF S II']
_range_chain = range(5, 28 + 1)
_range_chain = chain(_range_chain, range(30, 31+1))
_range_chain = chain(_range_chain, range(36, 38+1))
_range_chain = chain(_range_chain, range(40, 47+1))
_range_chain = chain(_range_chain, range(51, 53+1))

for _i in _range_chain:
    _third['E'+str(_i)].value = _third['G'+str(_i)].value

_third['G5'].value = MB_corr_number('Zinsrisiko', 'SCR netto')
_third['G6'].value = MB_corr_number('Aktienrisiko', 'SCR netto')
_third['G7'].value = MB_corr_number('Immobilienrisiko', 'SCR netto')
_third['G8'].value = MB_corr_number('Spreadrisiko', 'SCR netto')
_third['G9'].value = MB_corr_number('Marktrisikokonzentration', 'SCR netto')
_third['G10'].value = MB_corr_number('Wechselkursrisiko', 'SCR netto')
_third['G11'].value = MB_corr_number('Marktrisiko', 'SCR netto')
_third['G12'].value = MB_corr_number('Sterblichkeitsrisiko', 'SCR netto')
_third['G13'].value = MB_corr_number('Langlebigkeitsrisiko', 'SCR netto')
_third['G14'].value = MB_corr_number('Invaliditätsrisiko', 'SCR netto')
_third['G15'].value = MB_corr_number('Stornorisiko', 'SCR netto')
_third['G16'].value = MB_corr_number('Kostenrisiko', 'SCR netto')
_third['G17'].value = MB_corr_number('Revisionsrisiko', 'SCR netto')
_third['G18'].value = MB_corr_number('vt. Risiko Kranken SLT', 'SCR netto')
_third['G19'].value = MB_corr_number('Massenunfallrisiko', 'SCR netto')
_third['G20'].value = MB_corr_number('Pandemie', 'SCR netto')
_third['G21'].value = MB_corr_number('vt. Risiko Kranken CAT', 'SCR netto')
_third['G22'].value = MB_corr_number('vt. Risiko Kranken', 'SCR netto')
_third['G23'].value = MB_corr_number('Gegenparteiausfallrisiko Typ 1', 'SCR netto')
_third['G24'].value = MB_corr_number('Gegenparteiausfallrisiko Typ 2', 'SCR netto')
_third['G25'].value = MB_corr_number('Gegenparteiausfallrisiko', 'SCR netto')
_third['G26'].value = MB_corr_number('BSCR', 'SCR netto')
_third['G28'].value = MB_corr_number('OpRisk', 'SCR netto')
_third['G30'].value = MB_corr_number('AdjDT', 'SCR netto')
_third['G36'].value = MB_corr_number('Kapitalanlagen Direktbestand', 'SCR netto')
_third['G38'].value = MB_corr_number('sonstige Aktiva', 'SCR netto')
_third['G41'].value = MB_corr_number('davon Risikomarge', 'SCR netto')
_third['G42'].value = MB_corr_number('davon bester Schätzwert', 'SCR netto')
_third['G44'].value = MB_corr_number('passive latente Steuer', 'SCR netto')
_third['G45'].value = MB_corr_number('restliche Passiva', 'SCR netto')
_third['G47'].value = MB_corr_number('Eigenmittel', 'SCR netto')
_third['G51'].value = MB_corr_number('Tier 1', 'SCR netto')

IRB.save('IRB_Q3.xlsx')

#init class training

class Enemy:
    def __init__(self,x):
        self.energy = x

    def get_energy(self):
        print(self.energy)

    def attack(self):
        print('ouch')
        self.energy -=1

    def checkLife(self):
        if self.energy <=0:
            print('I am dead')
        else:
            print(str(self.energy)+" life left")

enemy1 = Enemy(1)
enemy2 = Enemy(18)

enemy1.get_energy()
enemy2.get_energy()

enemy1.attack()
enemy1.checkLife()
enemy2.attack()
enemy2.checkLife()
